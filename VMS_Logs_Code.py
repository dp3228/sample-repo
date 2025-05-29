import cv2
import pickle
import numpy as np
import threading
from ultralytics import YOLO
from deepface import DeepFace
from sklearn.metrics.pairwise import cosine_similarity
import json
import re
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Load YOLO model once
model = YOLO("C:\\Users\\dlnar\\Desktop\\Sparquer\\YOLOV8m_best.pt")

# Load face database
with open("face_database.pkl", "rb") as file:
    face_database = pickle.load(file)

# Load camera mapping
with open("camera_map.json", "r", encoding="utf-8") as f:
    camera_mapping = json.load(f)

# Load access permissions
with open("authorized_access.json", "r", encoding="utf-8") as f:
    authorized_access = json.load(f)

SIMILARITY_THRESHOLD = 0.3
EXCEL_LOG_PATH = "visitor_log_1.xlsx"

# Initialize Excel log file
if not os.path.exists(EXCEL_LOG_PATH):
    wb = Workbook()
    ws = wb.active
    ws.title = "Visitor Log"
    ws.append(["Time", "Date", "Status", "Person Name", "Camera Name"])
    wb.save(EXCEL_LOG_PATH)

def get_face_embedding(face_image):
    embedding = DeepFace.represent(face_image, model_name='ArcFace', enforce_detection=False)[0]['embedding']
    return embedding

def log_visitor(status, person_name, camera_name):
    time_str = datetime.now().strftime("%H:%M:%S")
    date_str = datetime.now().strftime("%Y-%m-%d")

    wb = load_workbook(EXCEL_LOG_PATH)
    ws = wb.active
    ws.append([time_str, date_str, status, person_name, camera_name])
    wb.save(EXCEL_LOG_PATH)

def process_camera(cam_url):
    # Extract port and map to camera name
    port_match = re.search(r":(\d+)", cam_url)
    if port_match:
        port = port_match.group(1)
        camera_name = camera_mapping.get(port, f"Unknown Port {port}")
    else:
        camera_name = "Unknown Camera"

    cap = cv2.VideoCapture(cam_url)
    if not cap.isOpened():
        print(f"[ERROR] Failed to open stream: {cam_url}")
        return

    print(f"[INFO] Started stream: {camera_name}")

    while True:
        ret, frame = cap.read()
        if not ret:
            print(f"[ERROR] Frame read failed from {camera_name}")
            break

        # Display camera name
        cv2.putText(frame, camera_name, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2)

        try:
            results = model.predict(frame)
            for result in results[0].boxes.data:
                x1, y1, x2, y2 = map(int, result[:4])
                conf = float(result[4])
                if conf < 0.4:
                    continue

                h, w = frame.shape[:2]
                x1, y1 = max(0, x1), max(0, y1)
                x2, y2 = min(w, x2), min(h, y2)

                face_image = frame[y1:y2, x1:x2]
                if face_image.size == 0 or (x2 - x1 < 20 or y2 - y1 < 20):
                    continue

                face_image = cv2.resize(face_image, (160, 160))
                embedding = get_face_embedding(face_image)

                best_label = "Unknown"
                best_similarity = 0.0

                for label, stored_embeddings in face_database.items():
                    similarities = cosine_similarity([embedding], stored_embeddings)
                    max_sim = max(similarities[0])
                    if max_sim > best_similarity:
                        best_similarity = max_sim
                        if best_similarity >= SIMILARITY_THRESHOLD:
                            best_label = label

                # Determine if access is authorized
                is_unauthorized = False
                if best_label != "Unknown":
                    allowed_locations = authorized_access.get(best_label, [])
                    if camera_name not in allowed_locations:
                        is_unauthorized = True
                        print(f"[ALERT] {best_label} spotted in unauthorized location: {camera_name}")

                color = (0, 0, 255) if is_unauthorized else (0, 255, 0)
                status = "Unauthorized" if is_unauthorized else "Authorized"
                person_name = best_label if best_label != "Unknown" else "Unknown"
                box_text = f"{best_label}"
                # ({best_similarity:.2f})

                if is_unauthorized:
                    box_text += " - UNAUTHORIZED"
                else: 
                    box_text += " - AUTHORIZED"
                
                if (best_similarity >= SIMILARITY_THRESHOLD):
                    cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                    cv2.putText(frame, box_text, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)
                
                log_visitor(status, person_name, camera_name)

        except Exception as e:
            print(f"[ERROR] in {camera_name}: {e}")

        cv2.imshow(camera_name, frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyWindow(camera_name)
    print(f"[INFO] Stream stopped: {camera_name}")

# --- Camera URLs ---
camera_urls = [
    #"rtsp://admin:admin123@202.53.64.30:1043/cam/realmonitor?channel=1&subtype=0",
    "rtsp://admin:admin123@202.53.64.30:1039/cam/realmonitor?channel=1&subtype=0"
]

# --- Start threads ---
threads = []
for url in camera_urls:
    t = threading.Thread(target=process_camera, args=(url,))
    t.start()
    threads.append(t)

# --- Wait for all threads ---
for t in threads:
    t.join()

cv2.destroyAllWindows()
